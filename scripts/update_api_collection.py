from copy import copy
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(r"E:\Ministry of Home Affairs - Completed API Collection.xlsx")
OUTPUT = Path(r"C:\Users\hirun\OneDrive\Desktop\Life Location Code Management System\outputs\Ministry of Home Affairs - Sanctum API Collection.xlsx")


NO_BODY = '{\n  "note": "No request body"\n}'
AUTH_NOTE = (
    "Requires Authorization: Bearer {access_token} with the location:read ability. "
    "Guest tokens expire after 60 minutes; officer and administrator tokens expire after 8 hours."
)


rows = [
    {
        "id": "LLCMS-API-001",
        "name": "Issue Guest Access Token",
        "description": "Issues a read-only Laravel Sanctum token for anonymous public website access.",
        "method": "POST",
        "url": "/api/v1/auth/guest-token",
        "parameters": "(none)",
        "request": NO_BODY,
        "response": '{\n  "token": "18|example-token",\n  "token_type": "Bearer",\n  "expires_at": "2026-08-10T05:26:58+00:00",\n  "abilities": ["location:read"]\n}',
        "sample": "POST /api/v1/auth/guest-token",
        "notes": "No authentication required. Rate limited to 10 requests per minute. Token expires after 60 minutes.",
    },
    {
        "id": "LLCMS-API-002",
        "name": "List Provinces",
        "description": "Returns all provinces ordered by English name, including English, Sinhala and Tamil names.",
        "method": "GET",
        "url": "/api/v1/locations/provinces",
        "parameters": "Authorization (header, required) - Bearer {access_token}",
        "request": NO_BODY,
        "response": '[\n  {\n    "id": 1,\n    "name_english": "Western",\n    "name_sinhala": "බස්නාහිර",\n    "name_tamil": "மேல்",\n    "province_code": "1",\n    "lifecode": "1"\n  }\n]',
        "sample": "GET /api/v1/locations/provinces",
        "notes": AUTH_NOTE,
    },
    {
        "id": "LLCMS-API-003",
        "name": "List Districts",
        "description": "Returns districts ordered by English name; optionally filters them by province.",
        "method": "GET",
        "url": "/api/v1/locations/districts",
        "parameters": "Authorization (header, required) - Bearer {access_token}\nprovince_id (query, integer, optional) - parent province ID",
        "request": NO_BODY,
        "response": '[\n  {\n    "id": 1,\n    "name_english": "Colombo",\n    "name_sinhala": "කොළඹ",\n    "name_tamil": "கொழும்பு",\n    "district_code": "1",\n    "lifecode": "1-1",\n    "province_id": 1\n  }\n]',
        "sample": "GET /api/v1/locations/districts?province_id=1",
        "notes": AUTH_NOTE + " Invalid province IDs return HTTP 422.",
    },
    {
        "id": "LLCMS-API-004",
        "name": "List Divisional Secretariats",
        "description": "Returns divisional secretariats ordered by English name; optionally filters by district.",
        "method": "GET",
        "url": "/api/v1/locations/divisional-secretariats",
        "parameters": "Authorization (header, required) - Bearer {access_token}\ndistrict_id (query, integer, optional) - parent district ID",
        "request": NO_BODY,
        "response": '[\n  {\n    "id": 12,\n    "name_english": "Colombo",\n    "name_sinhala": "කොළඹ",\n    "name_tamil": "கொழும்பு",\n    "divisional_secretariat_code": "01",\n    "lifecode": "1-1-01",\n    "district_id": 1\n  }\n]',
        "sample": "GET /api/v1/locations/divisional-secretariats?district_id=1",
        "notes": AUTH_NOTE + " Invalid district IDs return HTTP 422.",
    },
    {
        "id": "LLCMS-API-005",
        "name": "List GN Divisions",
        "description": "Returns Grama Niladhari divisions ordered by English name; optionally filters by DS division.",
        "method": "GET",
        "url": "/api/v1/locations/gn-divisions",
        "parameters": "Authorization (header, required) - Bearer {access_token}\nds_id (query, integer, optional) - parent divisional secretariat ID",
        "request": NO_BODY,
        "response": '[\n  {\n    "id": 101,\n    "name_english": "Example GN",\n    "name_sinhala": "උදාහරණ",\n    "name_tamil": "உதாரணம்",\n    "grama_niladhari_division_code": "001",\n    "lifecode": "1-1-01-001",\n    "mpa_code": "MPA-001",\n    "divisional_secretariat_id": 12\n  }\n]',
        "sample": "GET /api/v1/locations/gn-divisions?ds_id=12",
        "notes": AUTH_NOTE + " Invalid DS IDs return HTTP 422.",
    },
    {
        "id": "LLCMS-API-006",
        "name": "List Villages",
        "description": "Returns villages ordered by English name; optionally filters by GN division.",
        "method": "GET",
        "url": "/api/v1/locations/villages",
        "parameters": "Authorization (header, required) - Bearer {access_token}\ngn_id (query, integer, optional) - parent GN division ID",
        "request": NO_BODY,
        "response": '[\n  {\n    "id": 1001,\n    "name_english": "Example Village",\n    "name_sinhala": "උදාහරණ ගම",\n    "name_tamil": "உதாரண கிராமம்",\n    "village_code": "001",\n    "lifecode": "1-1-01-001-001",\n    "grama_niladhari_division_id": 101\n  }\n]',
        "sample": "GET /api/v1/locations/villages?gn_id=101",
        "notes": AUTH_NOTE + " Invalid GN IDs return HTTP 422.",
    },
    {
        "id": "LLCMS-API-007",
        "name": "Search Location Codes",
        "description": "Searches the province-to-village hierarchy and returns paginated records for the selected hierarchy level.",
        "method": "GET",
        "url": "/api/v1/search",
        "parameters": "Authorization (header, required) - Bearer {access_token}\nprovince_id, district_id, ds_id, gn_id (query, integer|'all'|'none', optional)\nkeyword (query, string, optional)\ninclude_villages (query, boolean, optional)\nsort_by (query, 'name'|'code', optional)\nper_page (query, integer, optional, max 100)\npage (query, integer, optional)",
        "request": NO_BODY,
        "response": '{\n  "current_page": 1,\n  "data": [\n    {\n      "village_id": 1001,\n      "village_name": "Example Village",\n      "village_lifecode": "1-1-01-001-001",\n      "gn_id": 101,\n      "gn_name": "Example GN",\n      "ds_id": 12,\n      "ds_name": "Colombo",\n      "district_id": 1,\n      "district_name": "Colombo",\n      "province_id": 1,\n      "province_name": "Western"\n    }\n  ],\n  "per_page": 25,\n  "total": 1\n}',
        "sample": "GET /api/v1/search?province_id=1&keyword=Colombo&per_page=25",
        "notes": AUTH_NOTE,
    },
    {
        "id": "LLCMS-API-008",
        "name": "Duplicate GN Analysis",
        "description": "Finds GN division names occurring under different divisional secretariats within the same district.",
        "method": "GET",
        "url": "/api/v1/duplicate-gn",
        "parameters": "Authorization (header, required) - Bearer {access_token}\nprovince_id (query, integer|'all', optional)\ndistrict_id (query, integer|'all', optional)",
        "request": NO_BODY,
        "response": '{\n  "data": [\n    {\n      "province_name": "Central",\n      "district_name": "Matale",\n      "ds_name": "Dambulla",\n      "gn_name": "Example",\n      "gn_lifecode": "2-2-06-001",\n      "province_id": 2,\n      "district_id": 2,\n      "ds_id": 31\n    }\n  ],\n  "summary": {\n    "total_rows": 1,\n    "province_count": 1,\n    "district_count": 1,\n    "ds_count": 1,\n    "gn_group_count": 1\n  }\n}',
        "sample": "GET /api/v1/duplicate-gn?province_id=2&district_id=2",
        "notes": AUTH_NOTE,
    },
]


def copy_cell_style(source, target):
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


wb = load_workbook(SOURCE)
ws = wb["API "]

# Rows 2-7 already contain API entries and rows below them are formatted blanks.
# Copy the established API row style to all eight populated rows.
for row_index in range(2, 10):
    for col_index in range(1, 26):
        copy_cell_style(ws.cell(2, col_index), ws.cell(row_index, col_index))
    ws.row_dimensions[row_index].height = ws.row_dimensions[2].height

for row_index, api in enumerate(rows, start=2):
    values = [
        api["id"], api["name"], api["description"], api["method"], api["url"],
        api["parameters"], api["request"], api["response"], api["sample"], api["notes"],
        True, True, "Real API", None, "Readily available", "Public",
    ]
    for col_index, value in enumerate(values, start=1):
        ws.cell(row_index, col_index).value = value

# Keep the workbook metadata aligned with the versioned API base.
info = wb["Info"]
info["B8"] = "http://127.0.0.1:8000/api/v1"

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(OUTPUT)
